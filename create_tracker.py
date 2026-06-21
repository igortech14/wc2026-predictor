import openpyxl
from openpyxl.styles import Font, Alignment, numbers

def create_tracker():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "World Cup Bets"

    # ---------- Headers ----------
    headers = ["Date", "Match", "Bet Type", "Odds", "Kelly %", "Stake Amount",
               "Result", "Profit/Loss (after tax)", "Bankroll"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # ---------- Starting bankroll ----------
    starting_bankroll = 18000
    ws.cell(row=1, column=9, value=starting_bankroll)   # I1
    ws.cell(row=1, column=9).font = Font(bold=True)

    # ---------- Formulas for rows 2 to 100 ----------
    for row in range(2, 101):
        # A, B, C, D, E are manually filled
        # F: Stake Amount = (Kelly % / 100) * Bankroll from previous row
        ws.cell(row=row, column=6).value = f"=E{row}*I{row-1}/100"
        ws.cell(row=row, column=6).number_format = '#,##0.00'

        # G: Result (manually entered as "Win" or "Loss")
        # H: Profit/Loss after tax
        # If win: profit = stake * (odds-1) * 0.85
        # If loss: profit = -stake
        ws.cell(row=row, column=8).value = \
            f'=IF(G{row}="Win", F{row}*(D{row}-1)*0.85, -F{row})'
        ws.cell(row=row, column=8).number_format = '#,##0.00'

        # I: Bankroll = previous bankroll + profit/loss
        ws.cell(row=row, column=9).value = f"=I{row-1}+H{row}"
        ws.cell(row=row, column=9).number_format = '#,##0.00'

    # ---------- Adjust column widths ----------
    widths = [12, 30, 15, 8, 8, 14, 8, 16, 12]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    # ---------- Add a summary area (optional) ----------
    ws.cell(row=103, column=1, value="Final Bankroll:").font = Font(bold=True)
    ws.cell(row=103, column=2, value="=LOOKUP(2,1/(I2:I100<>\"\"),I2:I100)")
    ws.cell(row=103, column=2).number_format = '#,##0.00'

    ws.cell(row=104, column=1, value="Total Profit:").font = Font(bold=True)
    ws.cell(row=104, column=2, value=f"=B103-{starting_bankroll}")
    ws.cell(row=104, column=2).number_format = '#,##0.00'

    wb.save("world_cup_tracker.xlsx")
    print("Tracker created: world_cup_tracker.xlsx")

if __name__ == "__main__":
    # Make sure openpyxl is installed: pip install openpyxl
    create_tracker()